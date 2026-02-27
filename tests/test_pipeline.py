"""Suite de pruebas del pipeline CxC — Microsip.

Ejecuta tres niveles de prueba en orden:

    Nivel 1 — Sin base de datos (datos sintéticos)
        Verifica que cada módulo de src/ procesa datos correctamente
        sin necesitar conexión a Firebird.  Siempre debe pasar.

    Nivel 2 — Conexión a Firebird
        Verifica que la configuración de red/credenciales es correcta
        y que el query SQL maestro devuelve datos con la estructura
        esperada.  Requiere que la PC pueda alcanzar el servidor.

    Nivel 3 — Pipeline completo end-to-end
        Corre run_pipeline() completo y verifica que el Excel se
        genera con todas las pestañas esperadas.

Uso:
    # Todos los niveles
    python tests/test_pipeline.py

    # Solo nivel 1 (sin DB, siempre disponible)
    python tests/test_pipeline.py --nivel 1

    # Niveles 1 y 2
    python tests/test_pipeline.py --nivel 2

    # Verbose: muestra detalles de cada prueba
    python tests/test_pipeline.py --verbose
"""

from __future__ import annotations

import argparse
import sys
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

# ── Asegurar que el raíz del proyecto esté en el path ─────────────────
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))


# ======================================================================
# COLORES Y UTILIDADES DE CONSOLA
# ======================================================================
class C:
    """Códigos ANSI para colores en terminal."""
    OK      = "\033[92m"   # Verde
    WARN    = "\033[93m"   # Amarillo
    FAIL    = "\033[91m"   # Rojo
    BOLD    = "\033[1m"
    RESET   = "\033[0m"
    CYAN    = "\033[96m"
    GREY    = "\033[90m"


def _ok(msg: str) -> None:
    print(f"  {C.OK}✅ PASS{C.RESET}  {msg}")


def _fail(msg: str, detalle: str = "") -> None:
    print(f"  {C.FAIL}❌ FAIL{C.RESET}  {msg}")
    if detalle:
        for linea in detalle.strip().splitlines():
            print(f"         {C.GREY}{linea}{C.RESET}")


def _warn(msg: str) -> None:
    print(f"  {C.WARN}⚠️  WARN{C.RESET}  {msg}")


def _header(titulo: str) -> None:
    ancho = 60
    print(f"\n{C.BOLD}{C.CYAN}{'═' * ancho}{C.RESET}")
    print(f"{C.BOLD}{C.CYAN}  {titulo}{C.RESET}")
    print(f"{C.BOLD}{C.CYAN}{'═' * ancho}{C.RESET}")


def _subheader(titulo: str) -> None:
    print(f"\n{C.BOLD}  ── {titulo} ──{C.RESET}")


# ======================================================================
# FIXTURE DE DATOS SINTÉTICOS
# ======================================================================

def _df_sintetico(n: int = 50) -> pd.DataFrame:
    """Genera un DataFrame sintético que imita la salida del query maestro.

    Cubre todos los casos edge que los módulos necesitan:
        - Movimientos C (cargo) y R (abono) vinculados.
        - Facturas vencidas y vigentes.
        - Un documento cancelado.
        - Un importe atípico (outlier).
        - Un registro sin nombre de cliente.

    Args:
        n: Número de cargos a generar.

    Returns:
        DataFrame con la misma estructura que el query Firebird.
    """
    rng   = np.random.default_rng(42)
    hoy   = datetime.now().date()

    clientes = [
        "EMPRESA ALPHA SA", "COMERCIAL BETA SC", "GRUPO GAMMA SRL",
        "DISTRIBUIDORA DELTA", "SERVICIOS EPSILON",
    ]
    vendedores = ["CARLOS LÓPEZ", "ANA MARTÍNEZ", "ROBERTO SÁNCHEZ"]
    conceptos  = ["FACTURA", "NOTA CARGO", "INTERESES"]

    # Generar cargos
    filas: list[dict[str, Any]] = []
    for i in range(1, n + 1):
        fecha_emision    = hoy - timedelta(days=int(rng.integers(1, 200)))
        dias_credito     = rng.choice([30, 60, 90])
        fecha_vencimiento = fecha_emision + timedelta(days=int(dias_credito))
        importe          = round(float(rng.uniform(500, 50_000)), 2)
        impuesto         = round(importe * 0.16, 2)

        filas.append({
            "DOCTO_CC_ID":          i,
            "DOCTO_CC_ACR_ID":      None,  # será llenado en abonos
            "FOLIO":                f"FAC-{i:04d}",
            "TIPO_IMPTE":           "C",
            "NATURALEZA_CONCEPTO":  "C",
            "CONCEPTO":             rng.choice(conceptos),
            "DESCRIPCION":          f"Venta período {fecha_emision.strftime('%b %Y')}",
            "NOMBRE_CLIENTE":       rng.choice(clientes),
            "CLIENTE_ID":           int(rng.integers(1, len(clientes) + 1)),
            "TIPO_CLIENTE":         rng.choice(["CONTADO", "CREDITO"]),
            "VENDEDOR":             rng.choice(vendedores),
            "FECHA_EMISION":        pd.Timestamp(fecha_emision),
            "FECHA_VENCIMIENTO":    pd.Timestamp(fecha_vencimiento),
            "IMPORTE":              importe,
            "IMPUESTO":             impuesto,
            "CARGOS":               importe + impuesto,
            "ABONOS":               0.0,
            "MONEDA":               "MXN",
            "CONDICIONES":          f"Crédito {dias_credito} días",
            "ESTATUS_CLIENTE":      "ACTIVO",
            "CANCELADO":            "N",
            "LIMITE_CREDITO":       round(float(rng.uniform(50_000, 300_000)), 2),
            "FECHA_HORA_CREACION":  pd.Timestamp(fecha_emision),
            "FECHA_HORA_ULT_MODIF": pd.Timestamp(fecha_emision),
            "FECHA_HORA_CANCELACION": None,
        })

    # Agregar abonos parciales (40% de los cargos tienen al menos un abono)
    abono_id = n + 1
    abonos: list[dict[str, Any]] = []
    for cargo in rng.choice(filas, size=int(n * 0.4), replace=False):
        abono_monto = round(float(cargo["IMPORTE"]) * float(rng.uniform(0.3, 1.0)), 2)
        abonos.append({
            **cargo,
            "DOCTO_CC_ID":     abono_id,
            "DOCTO_CC_ACR_ID": cargo["DOCTO_CC_ID"],
            "FOLIO":           f"REC-{abono_id:04d}",
            "TIPO_IMPTE":      "R",
            "NATURALEZA_CONCEPTO": "A",
            "IMPORTE":         abono_monto,
            "IMPUESTO":        round(abono_monto * 0.16, 2),
            "CARGOS":          0.0,
            "ABONOS":          abono_monto,
            "CANCELADO":       "N",
        })
        abono_id += 1

    df = pd.DataFrame(filas + abonos)

    # ── Casos edge especiales ──────────────────────────────────────────

    # 1. Documento cancelado
    df.loc[0, "CANCELADO"] = "S"
    df.loc[0, "FECHA_HORA_CANCELACION"] = pd.Timestamp(hoy)

    # 2. Importe atípico (outlier Z-score > 3)
    media  = df.loc[df["TIPO_IMPTE"] == "C", "IMPORTE"].mean()
    std    = df.loc[df["TIPO_IMPTE"] == "C", "IMPORTE"].std()
    df.loc[1, "IMPORTE"] = round(media + std * 4.5, 2)  # Z-score ~4.5

    # 3. Registro sin nombre de cliente
    df.loc[2, "NOMBRE_CLIENTE"] = None

    # 4. Duplicado intencional (copiar fila 3 con mismo CLIENTE_ID+FOLIO+CONCEPTO)
    fila_dup = df.loc[3].copy()
    fila_dup["DOCTO_CC_ID"] = abono_id + 10
    df = pd.concat([df, fila_dup.to_frame().T], ignore_index=True)

    return df


# ======================================================================
# NIVEL 1 — MÓDULOS CON DATOS SINTÉTICOS
# ======================================================================

class TestNivel1:
    """Pruebas de unidad/integración sin base de datos."""

    def __init__(self, verbose: bool = False) -> None:
        self.verbose = verbose
        self.passed  = 0
        self.failed  = 0
        self.df      = _df_sintetico()

    def _assert(self, condicion: bool, nombre: str, detalle: str = "") -> None:
        if condicion:
            _ok(nombre)
            self.passed += 1
        else:
            _fail(nombre, detalle)
            self.failed += 1

    # ------------------------------------------------------------------
    # SETTINGS
    # ------------------------------------------------------------------
    def test_settings(self) -> None:
        _subheader("config/settings.py")
        try:
            from config.settings import (
                ANOMALIAS, FIREBIRD_CONFIG, KPI_PERIODO_DIAS,
                OUTPUT_DIR, RANGOS_ANTIGUEDAD, SQL_FILE,
            )
            self._assert(isinstance(FIREBIRD_CONFIG, dict),     "FIREBIRD_CONFIG es dict")
            self._assert("host"     in FIREBIRD_CONFIG,          "FIREBIRD_CONFIG tiene 'host'")
            self._assert("database" in FIREBIRD_CONFIG,          "FIREBIRD_CONFIG tiene 'database'")
            self._assert("password" in FIREBIRD_CONFIG,          "FIREBIRD_CONFIG tiene 'password'")
            self._assert(isinstance(RANGOS_ANTIGUEDAD, list),    "RANGOS_ANTIGUEDAD es lista")
            self._assert(len(RANGOS_ANTIGUEDAD) > 0,             "RANGOS_ANTIGUEDAD no está vacía")
            self._assert(isinstance(ANOMALIAS, dict),            "ANOMALIAS es dict")
            self._assert("importe_zscore_umbral" in ANOMALIAS,   "ANOMALIAS tiene zscore_umbral")
            self._assert("dias_vencimiento_critico" in ANOMALIAS,"ANOMALIAS tiene dias_critico")
            self._assert(isinstance(KPI_PERIODO_DIAS, int),      "KPI_PERIODO_DIAS es entero")
            self._assert(SQL_FILE.suffix == ".sql",               "SQL_FILE apunta a archivo .sql")
        except Exception as e:
            _fail("Error importando settings", traceback.format_exc())
            self.failed += 1

    # ------------------------------------------------------------------
    # REPORTE CXC
    # ------------------------------------------------------------------
    def test_reporte_cxc(self) -> None:
        _subheader("src/reporte_cxc.py")
        try:
            from src.reporte_cxc import generar_reporte_cxc

            resultado = generar_reporte_cxc(self.df)

            self._assert(isinstance(resultado, dict),                   "generar_reporte_cxc devuelve dict")
            self._assert("reporte_cxc"    in resultado,                 "Clave 'reporte_cxc' presente")
            self._assert("por_acreditar"  in resultado,                 "Clave 'por_acreditar' presente")
            self._assert("facturas_vivas" in resultado,                 "Clave 'facturas_vivas' presente")

            reporte = resultado["reporte_cxc"]
            self._assert(isinstance(reporte, pd.DataFrame),             "reporte_cxc es DataFrame")
            self._assert(len(reporte) > 0,                              "reporte_cxc tiene filas")
            self._assert("SALDO_FACTURA" in reporte.columns,            "Columna SALDO_FACTURA existe")
            self._assert("SALDO_CLIENTE" in reporte.columns,            "Columna SALDO_CLIENTE existe")
            self._assert("CATEGORIA_MORA" in reporte.columns,           "Columna CATEGORIA_MORA existe")
            self._assert("DELTA_MORA" in reporte.columns,               "Columna DELTA_MORA existe")

            # Cancelados excluidos
            if "CANCELADO" in reporte.columns:
                cancelados_en_reporte = reporte[reporte["CANCELADO"].isin(["S", "SI"])].shape[0]
                self._assert(cancelados_en_reporte == 0,                "Cancelados excluidos del reporte")

            # Saldos numéricos
            self._assert(
                pd.to_numeric(reporte["SALDO_FACTURA"], errors="coerce").notna().any(),
                "SALDO_FACTURA contiene valores numéricos válidos",
            )

            # Facturas vivas
            fv = resultado["facturas_vivas"]
            if not fv.empty:
                self._assert("_BAND_GROUP" in fv.columns,               "facturas_vivas tiene _BAND_GROUP")
                self._assert(fv["_BAND_GROUP"].isin([0, 1]).all(),       "_BAND_GROUP solo tiene 0 o 1")

        except Exception as e:
            _fail("Error en reporte_cxc", traceback.format_exc())
            self.failed += 1

    # ------------------------------------------------------------------
    # ANALYTICS
    # ------------------------------------------------------------------
    def test_analytics(self) -> None:
        _subheader("src/analytics.py")
        try:
            from config.settings import RANGOS_ANTIGUEDAD
            from src.analytics import Analytics

            analytics  = Analytics(RANGOS_ANTIGUEDAD)
            resultados = analytics.run_analytics(self.df)

            claves_esperadas = [
                "antiguedad_cartera",
                "antiguedad_por_cliente",
                "cartera_vencida_vs_vigente",
                "resumen_por_cliente",
                "resumen_por_vendedor",
                "resumen_por_concepto",
                "datos_completos",
            ]
            for clave in claves_esperadas:
                self._assert(clave in resultados, f"Clave '{clave}' presente en analytics")

            ant = resultados["antiguedad_cartera"]
            self._assert(isinstance(ant, pd.DataFrame),                  "antiguedad_cartera es DataFrame")
            self._assert("IMPORTE_TOTAL" in ant.columns,                  "Columna IMPORTE_TOTAL presente")
            self._assert("PCT_DEL_TOTAL" in ant.columns,                  "Columna PCT_DEL_TOTAL presente")

            if not ant.empty:
                pct_suma = round(ant["PCT_DEL_TOTAL"].sum(), 1)
                self._assert(
                    abs(pct_suma - 100.0) < 1.0,
                    f"PCT_DEL_TOTAL suma ~100% (actual: {pct_suma}%)",
                )

            vv = resultados["cartera_vencida_vs_vigente"]
            if not vv.empty:
                self._assert(
                    "ESTATUS_VENCIMIENTO" in vv.columns,
                    "cartera_vencida_vs_vigente tiene ESTATUS_VENCIMIENTO",
                )

            rc = resultados["resumen_por_cliente"]
            self._assert(isinstance(rc, pd.DataFrame),                   "resumen_por_cliente es DataFrame")
            self._assert(len(rc) > 0,                                    "resumen_por_cliente tiene filas")

        except Exception as e:
            _fail("Error en analytics", traceback.format_exc())
            self.failed += 1

    # ------------------------------------------------------------------
    # AUDITOR
    # ------------------------------------------------------------------
    def test_auditor(self) -> None:
        _subheader("src/auditor.py")
        try:
            from config.settings import ANOMALIAS
            from src.auditor import Auditor, AuditResult

            auditor = Auditor(ANOMALIAS)
            result  = auditor.run_audit(self.df)

            self._assert(isinstance(result, AuditResult),                "run_audit devuelve AuditResult")
            self._assert(isinstance(result.resumen, dict),               "resumen es dict")
            self._assert("total_registros" in result.resumen,            "resumen tiene total_registros")
            self._assert("total_hallazgos" in result.resumen,            "resumen tiene total_hallazgos")

            # Debe detectar el cancelado que pusimos
            self._assert(
                len(result.documentos_cancelados) >= 1,
                f"Detectó al menos 1 cancelado (encontró {len(result.documentos_cancelados)})",
            )

            # Debe detectar el importe atípico que pusimos
            self._assert(
                len(result.importes_atipicos) >= 1,
                f"Detectó al menos 1 importe atípico (encontró {len(result.importes_atipicos)})",
            )

            # Debe detectar el duplicado que pusimos
            self._assert(
                len(result.duplicados) >= 2,
                f"Detectó registros duplicados (encontró {len(result.duplicados)})",
            )

            # Debe detectar el registro sin nombre de cliente
            self._assert(
                len(result.registros_sin_cliente) >= 1,
                f"Detectó registro sin cliente (encontró {len(result.registros_sin_cliente)})",
            )

            # Calidad de datos
            self._assert(isinstance(result.calidad_datos, pd.DataFrame), "calidad_datos es DataFrame")
            self._assert(len(result.calidad_datos) > 0,                  "calidad_datos tiene filas")
            self._assert("PCT_NULOS" in result.calidad_datos.columns,    "calidad_datos tiene PCT_NULOS")

        except Exception as e:
            _fail("Error en auditor", traceback.format_exc())
            self.failed += 1

    # ------------------------------------------------------------------
    # KPIS
    # ------------------------------------------------------------------
    def test_kpis(self) -> None:
        _subheader("src/kpis.py")
        try:
            from config.settings import KPI_PERIODO_DIAS
            from src.kpis import generar_kpis

            resultado = generar_kpis(self.df, KPI_PERIODO_DIAS)

            claves_esperadas = [
                "kpis_resumen",
                "kpis_concentracion",
                "kpis_limite_credito",
                "kpis_morosidad_cliente",
            ]
            for clave in claves_esperadas:
                self._assert(clave in resultado, f"Clave '{clave}' presente en KPIs")

            resumen = resultado["kpis_resumen"]
            self._assert(isinstance(resumen, pd.DataFrame),              "kpis_resumen es DataFrame")
            self._assert(len(resumen) == 3,                              "kpis_resumen tiene 3 filas (DSO, CEI, Morosidad)")
            self._assert("KPI" in resumen.columns,                       "Columna KPI presente")
            self._assert("VALOR" in resumen.columns,                     "Columna VALOR presente")
            self._assert("INTERPRETACION" in resumen.columns,            "Columna INTERPRETACION presente")

            # Verificar rangos lógicos de los KPIs
            kpis_dict = dict(zip(resumen["KPI"].str[:3], resumen["VALOR"]))
            dso_val = float(resumen[resumen["KPI"].str.contains("DSO")]["VALOR"].iloc[0])
            cei_val = float(resumen[resumen["KPI"].str.contains("CEI")]["VALOR"].iloc[0])
            mor_val = float(resumen[resumen["KPI"].str.contains("Morosidad")]["VALOR"].iloc[0])

            self._assert(dso_val >= 0,                                   f"DSO ≥ 0 (valor: {dso_val:.1f})")
            self._assert(0 <= cei_val <= 100,                            f"CEI en rango 0-100% (valor: {cei_val:.1f}%)")
            self._assert(0 <= mor_val <= 100,                            f"Morosidad en rango 0-100% (valor: {mor_val:.1f}%)")

            conc = resultado["kpis_concentracion"]
            if not conc.empty:
                self._assert("CLASIFICACION" in conc.columns,            "Concentración tiene CLASIFICACION ABC")
                self._assert(
                    conc["CLASIFICACION"].isin(["A", "B", "C"]).all(),
                    "Clasificación ABC solo contiene A, B o C",
                )
                self._assert(
                    conc["PCT_ACUMULADO"].max() <= 100.1,
                    f"PCT_ACUMULADO máximo ≤ 100 (valor: {conc['PCT_ACUMULADO'].max():.1f}%)",
                )

        except Exception as e:
            _fail("Error en kpis", traceback.format_exc())
            self.failed += 1

    # ------------------------------------------------------------------
    # EXPORTACIÓN EXCEL
    # ------------------------------------------------------------------
    def test_exportacion_excel(self) -> None:
        _subheader("main.py — exportar_a_excel()")
        try:
            import tempfile
            from main import exportar_a_excel
            from src.reporte_cxc import generar_reporte_cxc
            from src.analytics import Analytics
            from src.auditor import Auditor
            from src.kpis import generar_kpis
            from config.settings import RANGOS_ANTIGUEDAD, ANOMALIAS, KPI_PERIODO_DIAS

            # Generar todos los DataFrames
            reporte   = generar_reporte_cxc(self.df)
            analytics = Analytics(RANGOS_ANTIGUEDAD).run_analytics(self.df)
            audit     = Auditor(ANOMALIAS).run_audit(self.df)
            kpis      = generar_kpis(self.df, KPI_PERIODO_DIAS)

            all_outputs = {**reporte, **analytics, **kpis}
            all_outputs["audit_resumen"] = pd.DataFrame(
                [audit.resumen]
            ).T.reset_index().rename(columns={"index": "METRICA", 0: "VALOR"})

            # Exportar a directorio temporal
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp_path = Path(tmpdir)
                filepath = exportar_a_excel(all_outputs, "test_export", tmp_path)

                self._assert(filepath.exists(),                          "Archivo Excel creado")
                self._assert(filepath.suffix == ".xlsx",                 "Extensión .xlsx correcta")
                self._assert(filepath.stat().st_size > 0,               "Archivo Excel no está vacío")

                # Verificar pestañas
                import openpyxl
                wb = openpyxl.load_workbook(filepath)
                pestanas = wb.sheetnames

                pestanas_esperadas = ["reporte_cxc", "facturas_vivas", "por_acreditar"]
                for p in pestanas_esperadas:
                    self._assert(p in pestanas, f"Pestaña '{p}' existe en Excel")

                self._assert(len(pestanas) >= 3,                         f"Excel tiene al menos 3 pestañas ({len(pestanas)} encontradas)")

                # Verificar que reporte_cxc tiene encabezado
                ws = wb["reporte_cxc"]
                self._assert(ws.max_row > 1,                             "reporte_cxc tiene filas de datos")
                self._assert(ws.max_column > 1,                          "reporte_cxc tiene múltiples columnas")

                wb.close()

        except Exception as e:
            _fail("Error en exportación Excel", traceback.format_exc())
            self.failed += 1

    def run(self) -> tuple[int, int]:
        _header("NIVEL 1 — Módulos con datos sintéticos (sin DB)")
        self.test_settings()
        self.test_reporte_cxc()
        self.test_analytics()
        self.test_auditor()
        self.test_kpis()
        self.test_exportacion_excel()
        return self.passed, self.failed


# ======================================================================
# NIVEL 2 — CONEXIÓN A FIREBIRD
# ======================================================================

class TestNivel2:
    """Pruebas que requieren conexión real a Firebird."""

    def __init__(self, verbose: bool = False) -> None:
        self.verbose = verbose
        self.passed  = 0
        self.failed  = 0

    def _assert(self, condicion: bool, nombre: str, detalle: str = "") -> None:
        if condicion:
            _ok(nombre)
            self.passed += 1
        else:
            _fail(nombre, detalle)
            self.failed += 1

    def test_conexion(self) -> bool:
        """Verifica que la conexión a Firebird funciona.

        Returns:
            True si la conexión fue exitosa.
        """
        _subheader("Conexión a Firebird")
        try:
            from config.settings import FIREBIRD_CONFIG
            from src.db_connector import FirebirdConnector

            connector = FirebirdConnector(FIREBIRD_CONFIG)
            ok = connector.test_connection()
            self._assert(ok, f"Conexión a {FIREBIRD_CONFIG['database']}")
            return ok
        except ImportError as e:
            _fail("Driver de Firebird no instalado", str(e))
            _warn("Instala el driver: pip install fdb   (Firebird 2.5)")
            self.failed += 1
            return False
        except Exception as e:
            _fail("Error de conexión", traceback.format_exc())
            self.failed += 1
            return False

    def test_query_maestro(self) -> bool:
        """Ejecuta el SQL maestro y verifica la estructura del resultado.

        Returns:
            True si el query devolvió datos con las columnas esperadas.
        """
        _subheader("Query maestro SQL")
        try:
            from config.settings import FIREBIRD_CONFIG, SQL_FILE
            from src.db_connector import FirebirdConnector

            # Verificar que el archivo SQL existe
            if not SQL_FILE.exists():
                _fail(f"Archivo SQL no encontrado: {SQL_FILE}")
                self.failed += 1
                return False
            _ok(f"Archivo SQL encontrado: {SQL_FILE.name}")
            self.passed += 1

            connector = FirebirdConnector(FIREBIRD_CONFIG)
            df = connector.execute_sql_file(SQL_FILE)

            self._assert(isinstance(df, pd.DataFrame),                   "Query devuelve DataFrame")
            self._assert(len(df) > 0,                                    f"Query devuelve datos ({len(df):,} filas)")
            self._assert(len(df.columns) >= 10,                          f"Query devuelve suficientes columnas ({len(df.columns)})")

            # Columnas mínimas requeridas
            df.columns = pd.Index([c.upper().strip() for c in df.columns])
            columnas_requeridas = [
                "DOCTO_CC_ID", "TIPO_IMPTE", "IMPORTE",
                "FECHA_EMISION", "NOMBRE_CLIENTE",
            ]
            for col in columnas_requeridas:
                self._assert(col in df.columns, f"Columna requerida '{col}' presente")

            # Columnas recomendadas (warning si faltan)
            columnas_recomendadas = [
                "DOCTO_CC_ACR_ID", "FECHA_VENCIMIENTO", "CANCELADO",
                "VENDEDOR", "FOLIO", "IMPUESTO",
            ]
            for col in columnas_recomendadas:
                if col not in df.columns:
                    _warn(f"Columna recomendada '{col}' ausente — algunos KPIs estarán incompletos")

            # TIPO_IMPTE debe tener C y/o R
            tipos = df["TIPO_IMPTE"].astype(str).str.strip().str.upper().unique().tolist()
            self._assert(
                "C" in tipos,
                f"TIPO_IMPTE contiene cargos 'C' (tipos encontrados: {tipos})",
            )

            return True

        except Exception as e:
            _fail("Error ejecutando query maestro", traceback.format_exc())
            self.failed += 1
            return False

    def test_pipeline_con_datos_reales(self) -> None:
        """Corre todos los módulos con los datos reales de Firebird."""
        _subheader("Pipeline completo con datos reales")
        try:
            from config.settings import (
                ANOMALIAS, FIREBIRD_CONFIG, KPI_PERIODO_DIAS,
                RANGOS_ANTIGUEDAD, SQL_FILE,
            )
            from src.analytics import Analytics
            from src.auditor import Auditor
            from src.db_connector import FirebirdConnector
            from src.kpis import generar_kpis
            from src.reporte_cxc import generar_reporte_cxc

            connector = FirebirdConnector(FIREBIRD_CONFIG)
            df = connector.execute_sql_file(SQL_FILE)

            # Reporte CxC
            reporte = generar_reporte_cxc(df)
            self._assert("reporte_cxc" in reporte,                       "reporte_cxc generado con datos reales")

            n_clientes = 0
            if "NOMBRE_CLIENTE" in reporte["reporte_cxc"].columns:
                n_clientes = reporte["reporte_cxc"]["NOMBRE_CLIENTE"].nunique()
            _ok(f"  → {len(reporte['reporte_cxc']):,} movimientos, {n_clientes} clientes")
            self.passed += 1

            # Analytics
            analytics = Analytics(RANGOS_ANTIGUEDAD).run_analytics(df)
            self._assert("antiguedad_cartera" in analytics,              "Analytics generado con datos reales")

            # Auditor
            audit = Auditor(ANOMALIAS).run_audit(df)
            self._assert(audit.resumen.get("total_registros", 0) > 0,   "Auditor procesó datos reales")
            _ok(f"  → {audit.resumen.get('total_hallazgos', 0)} hallazgos en {audit.resumen.get('total_registros', 0):,} registros")
            self.passed += 1

            # KPIs
            kpis = generar_kpis(df, KPI_PERIODO_DIAS)
            self._assert("kpis_resumen" in kpis,                         "KPIs calculados con datos reales")

            if not kpis["kpis_resumen"].empty:
                for _, row in kpis["kpis_resumen"].iterrows():
                    _ok(f"  → {row['KPI']}: {row['VALOR']} {row['UNIDAD']}")
                    self.passed += 1

        except Exception as e:
            _fail("Error en pipeline con datos reales", traceback.format_exc())
            self.failed += 1

    def run(self) -> tuple[int, int]:
        _header("NIVEL 2 — Conexión real a Firebird")
        conexion_ok = self.test_conexion()
        if not conexion_ok:
            _warn("Saltando pruebas de nivel 2 — sin conexión a Firebird.")
            return self.passed, self.failed

        query_ok = self.test_query_maestro()
        if query_ok:
            self.test_pipeline_con_datos_reales()

        return self.passed, self.failed


# ======================================================================
# NIVEL 3 — PIPELINE COMPLETO END-TO-END
# ======================================================================

class TestNivel3:
    """Prueba del pipeline completo incluyendo exportación a disco."""

    def __init__(self, verbose: bool = False) -> None:
        self.verbose = verbose
        self.passed  = 0
        self.failed  = 0

    def _assert(self, condicion: bool, nombre: str, detalle: str = "") -> None:
        if condicion:
            _ok(nombre)
            self.passed += 1
        else:
            _fail(nombre, detalle)
            self.failed += 1

    def test_run_pipeline(self) -> None:
        _subheader("run_pipeline() end-to-end")
        try:
            from main import run_pipeline

            codigo_salida = run_pipeline(
                skip_audit=False,
                skip_analytics=False,
                skip_kpis=False,
            )
            self._assert(codigo_salida == 0, f"run_pipeline() devolvió 0 (éxito)")

            # Verificar que el archivo Excel fue creado
            from config.settings import OUTPUT_DIR
            archivos_xlsx = list(OUTPUT_DIR.glob("auditoria_cxc_*.xlsx"))
            self._assert(len(archivos_xlsx) > 0, "Archivo Excel generado en output/")

            if archivos_xlsx:
                ultimo = max(archivos_xlsx, key=lambda p: p.stat().st_mtime)
                tamaño_kb = ultimo.stat().st_size / 1024
                self._assert(tamaño_kb > 10, f"Excel tiene tamaño razonable ({tamaño_kb:.0f} KB)")
                _ok(f"  → Archivo: {ultimo.name} ({tamaño_kb:.0f} KB)")
                self.passed += 1

        except Exception as e:
            _fail("Error en run_pipeline()", traceback.format_exc())
            self.failed += 1

    def run(self) -> tuple[int, int]:
        _header("NIVEL 3 — Pipeline completo end-to-end")
        self.test_run_pipeline()
        return self.passed, self.failed


# ======================================================================
# RUNNER PRINCIPAL
# ======================================================================

def _resumen_final(total_pass: int, total_fail: int) -> None:
    """Imprime el resumen final con colores."""
    total = total_pass + total_fail
    _header("RESUMEN FINAL")

    if total_fail == 0:
        print(f"\n  {C.OK}{C.BOLD}🎉 TODAS LAS PRUEBAS PASARON{C.RESET}")
        print(f"  {C.OK}{total_pass}/{total} pruebas exitosas{C.RESET}\n")
    else:
        pct = (total_pass / total * 100) if total > 0 else 0
        print(f"\n  {C.BOLD}Resultados: {C.OK}{total_pass} PASS{C.RESET} | {C.FAIL}{total_fail} FAIL{C.RESET} de {total} pruebas ({pct:.0f}%){C.RESET}")
        if total_fail > 0:
            print(f"\n  {C.WARN}Revisa los errores arriba marcados con ❌ para diagnosticar el problema.{C.RESET}")
        print()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Suite de pruebas del pipeline CxC — Microsip",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--nivel", type=int, default=3, choices=[1, 2, 3],
        help="Nivel máximo de pruebas a ejecutar (1=sintético, 2=+DB, 3=+end-to-end). Default: 3",
    )
    parser.add_argument(
        "--verbose", action="store_true",
        help="Mostrar detalle completo de errores.",
    )
    args = parser.parse_args()

    print(f"\n{C.BOLD}Pipeline CxC — Suite de Pruebas{C.RESET}")
    print(f"{C.GREY}Iniciando: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}{C.RESET}")
    print(f"{C.GREY}Raíz del proyecto: {ROOT}{C.RESET}")
    print(f"{C.GREY}Nivel máximo: {args.nivel}{C.RESET}")

    total_pass = 0
    total_fail = 0

    # Nivel 1 — siempre se ejecuta
    t1 = TestNivel1(verbose=args.verbose)
    p, f = t1.run()
    total_pass += p
    total_fail += f

    # Nivel 2 — requiere Firebird
    if args.nivel >= 2:
        t2 = TestNivel2(verbose=args.verbose)
        p, f = t2.run()
        total_pass += p
        total_fail += f

    # Nivel 3 — end-to-end completo
    if args.nivel >= 3:
        t3 = TestNivel3(verbose=args.verbose)
        p, f = t3.run()
        total_pass += p
        total_fail += f

    _resumen_final(total_pass, total_fail)
    return 0 if total_fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())