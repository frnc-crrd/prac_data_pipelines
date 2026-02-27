"""Pipeline principal de auditoria CxC para Microsip.

Orquesta la extraccion de datos desde Firebird, generacion del reporte
operativo, auditoria de anomalias, analisis de cartera, KPIs estrategicos
y exportacion a cuatro archivos Excel independientes.

Los cuatro archivos generados son:
    01_cxc_TIMESTAMP.xlsx
    02_auditoria_TIMESTAMP.xlsx
    03_analisis_TIMESTAMP.xlsx
    04_kpis_TIMESTAMP.xlsx

Estructura de 01_cxc:
    movimientos_abiertos_cxc   Cargos con saldo pendiente + abonos parciales
    movimientos_cerrados_cxc   Cargos completamente cobrados + sus abonos
    movimientos_totales_cxc    Todas las columnas + Z-scores intercalados
    registros_por_acreditar_cxc Anticipos sin aplicar (filtro de totales)
    registros_cancelados_cxc    Documentos cancelados (filtro de totales)
    registros_totales_cxc       Todos los registros crudos agrupados
                                (protegida con contrasena)

Uso:
    python main.py                    # Pipeline completo
    python main.py --test-connection  # Solo probar conexion a Firebird
    python main.py --skip-audit       # Saltar auditoria de anomalias
    python main.py --skip-analytics   # Saltar analisis de cartera
    python main.py --skip-kpis        # Saltar KPIs estrategicos
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime, time as dt_time
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))

from config.settings import (
    ANOMALIAS,
    EXCEL_ENGINE,
    EXCEL_NOMBRES,
    FIREBIRD_CONFIG,
    KPI_PERIODO_DIAS,
    OUTPUT_DIR,
    RANGOS_ANTIGUEDAD,
    SHEET_PASSWORDS,
    SQL_FILE,
)
from src.analytics import Analytics
from src.auditor import Auditor
from src.db_connector import FirebirdConnector
from src.kpis import generar_kpis
from src.reporte_cxc import agregar_bandas_grupo, generar_reporte_cxc
from src.reporte_pdf import generar_reporte_pdf

# ======================================================================
# LOGGING
# ======================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("main")

# ======================================================================
# CONSTANTES DE FORMATO
# ======================================================================

COLUMNAS_MONEDA: set[str] = {
    "CARGOS", "ABONOS", "IMPORTE", "IMPUESTO",
    "SALDO_FACTURA", "SALDO_CLIENTE",
    "IMPORTE_TOTAL", "IMPORTE_PROMEDIO", "IMPORTE_MAX",
    "TOTAL_CARGOS", "TOTAL_ABONOS", "SALDO",
    "MONTO_CARGO", "MONTO_ABONOS", "DISPONIBLE",
    "SALDO_TOTAL", "SALDO_VIGENTE", "SALDO_VENCIDO",
    "LIMITE_CREDITO",
    # Columnas de analisis de cartera
    "SALDO_PENDIENTE", "FACTURAS_PAGADAS", "FACTURAS_VIGENTES",
    "IMPUESTO_TOTAL", "MONTO_TOTAL",
}

# Prefijos de columnas que deben recibir formato de moneda aunque tengan
# nombres dinamicos con rangos entre parentesis (p.ej. FACTURAS_VENCIDAS (0-30)).
_COLUMNAS_MONEDA_PREFIJOS: tuple[str, ...] = ("FACTURAS_VENCIDAS",)

COLUMNAS_FECHA: set[str] = {
    "FECHA_EMISION", "FECHA_VENCIMIENTO",
    "FECHA_HORA_CREACION", "FECHA_HORA_ULT_MODIF",
    "FECHA_HORA_CANCELACION",
}

# Columnas que se formatean como porcentaje con dos decimales en Excel.
# Se usa el formato "0.00%" que multiplica el valor por 100 al mostrarlo,
# por lo que los valores deben estar almacenados como fraccion (0-1).
# Si los valores estan en escala 0-100 se usa el formato '#,##0.00"%"'
# para mostrar el simbolo sin multiplicar.
COLUMNAS_PORCENTAJE: set[str] = {"PCT_DEL_TOTAL"}

# Hojas que se protegen contra edicion.
# La contrasena de cada hoja se lee de SHEET_PASSWORDS en settings.
# Si una hoja esta aqui pero no en SHEET_PASSWORDS, se protege
# con contrasena vacia (Excel solicita dejar el campo en blanco).
PESTANAS_PROTEGIDAS: set[str] = {"registros_totales_cxc"}

# Valores que Microsip usa para marcar un documento como cancelado.
_CANCELADO_VALUES: list[Any] = ["S", "SI", "s", "si", 1, True, "1"]

_HEADER_FONT: Font = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL: PatternFill = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid",
)
_HEADER_ALIGNMENT: Alignment = Alignment(
    horizontal="center", vertical="center",
)
_THIN_BORDER: Border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
_BAND_FILL: PatternFill = PatternFill(
    start_color="D9E2F3", end_color="D9E2F3", fill_type="solid",
)
_WHITE_FILL: PatternFill = PatternFill(
    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid",
)

# Encabezado verde para columnas calculadas por el pipeline.
# Complementa el azul (4472C4) del encabezado estandar y permite
# identificar de un vistazo que columna viene del query vs del analisis.
_CALC_HEADER_FILL: PatternFill = PatternFill(
    start_color="548235", end_color="548235", fill_type="solid",
)

# Columnas calculadas por el pipeline que reciben encabezado verde
# en la hoja movimientos_totales_cxc.
COLUMNAS_CALCULADAS_CXC: set[str] = {
    "SALDO_FACTURA",
    "SALDO_CLIENTE",
    "DELTA_RECAUDO",
    "ZSCORE_DELTA_RECAUDO",
    "ATIPICO_DELTA_RECAUDO",
    "CATEGORIA_RECAUDO",
    "DELTA_MORA",
    "ZSCORE_DELTA_MORA",
    "ATIPICO_DELTA_MORA",
    "CATEGORIA_MORA",
    "ZSCORE_IMPORTE",
    "ATIPICO_IMPORTE",
}

# ======================================================================
# PREPARACION DE DATOS
# ======================================================================


def _formatear_hora(valor: Any) -> str:
    """Convierte un valor de hora Firebird (datetime.time) a HH:MM:SS.

    Args:
        valor: Valor de la columna HORA (datetime.time, None o NaN).

    Returns:
        Cadena HH:MM:SS o cadena vacia si el valor es nulo.
    """
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    if isinstance(valor, dt_time):
        return valor.strftime("%H:%M:%S")
    if hasattr(valor, "strftime"):
        return valor.strftime("%H:%M:%S")
    return str(valor)


def _normalizar_fechas_hora(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza fechas como datetime64 y columna HORA como cadena.

    Operacion base comun a todas las vistas que se exportan a Excel.
    Las fechas como datetime64 permiten que Excel las reconozca para
    agrupar por anio/mes/dia sin conversion manual.

    Args:
        df: DataFrame directo del query SQL.

    Returns:
        Copia del DataFrame con fechas y HORA normalizadas.
    """
    df = df.copy()
    for col in ["FECHA_EMISION", "FECHA_VENCIMIENTO"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    if "HORA" in df.columns:
        df["HORA"] = df["HORA"].apply(_formatear_hora)
    return df


def preparar_registros_totales(df: pd.DataFrame) -> pd.DataFrame:
    """Prepara todos los registros crudos con bandas de color por grupo.

    Genera la vista completa sin ningun filtro: incluye cargos, abonos,
    cancelados y tipo A. Es la fuente de la que se derivan los filtros
    ``registros_por_acreditar_cxc`` y ``registros_cancelados``.

    El coloreo alterno por grupo (cargo + sus abonos) facilita la
    lectura visual en la hoja ``registros_totales_cxc``.

    Args:
        df: DataFrame directo del query SQL con tipos nativos de Firebird.

    Returns:
        DataFrame normalizado con columna ``_BAND_GROUP`` (0 o 1).
    """
    return agregar_bandas_grupo(_normalizar_fechas_hora(df))


def _filtrar_por_acreditar(df_totales: pd.DataFrame) -> pd.DataFrame:
    """Extrae los registros con TIPO_IMPTE = 'A' de los totales.

    Los registros tipo A son anticipos o pagos que no han sido aplicados
    a ninguna factura especifica. Al derivarse de ``registros_totales_cxc``
    heredan el mismo formato de agrupacion y ``_BAND_GROUP``.

    Args:
        df_totales: DataFrame ya preparado por ``preparar_registros_totales``.

    Returns:
        DataFrame filtrado con registros tipo A, o vacio si no hay ninguno.
    """
    if "TIPO_IMPTE" not in df_totales.columns:
        return pd.DataFrame()

    tipo_norm = (
        df_totales["TIPO_IMPTE"].astype(str).str.strip().str.upper()
    )
    mask_tipo_a = tipo_norm == "A"

    # Excluir cancelados para que no haya solapamiento con
    # registros_cancelados_cxc. Un registro tipo A cancelado
    # debe aparecer solo en la pestana de cancelados.
    if "CANCELADO" in df_totales.columns:
        mask_activos = ~df_totales["CANCELADO"].isin(_CANCELADO_VALUES)
        resultado = df_totales[mask_tipo_a & mask_activos].copy()
    else:
        resultado = df_totales[mask_tipo_a].copy()
    logger.info("Registros por acreditar: %d filas.", len(resultado))
    return resultado


def _filtrar_cancelados(df_totales: pd.DataFrame) -> pd.DataFrame:
    """Extrae los registros cancelados de los totales.

    Al derivarse de ``registros_totales_cxc`` heredan el mismo formato
    de agrupacion y ``_BAND_GROUP``.

    Args:
        df_totales: DataFrame ya preparado por ``preparar_registros_totales``.

    Returns:
        DataFrame filtrado con registros cancelados, o vacio si no hay.
    """
    if "CANCELADO" not in df_totales.columns:
        return pd.DataFrame()

    resultado = df_totales[
        df_totales["CANCELADO"].isin(_CANCELADO_VALUES)
    ].copy()
    logger.info("Registros cancelados: %d filas.", len(resultado))
    return resultado


# ======================================================================
# FORMATO EXCEL — FUNCIONES INTERNAS
# ======================================================================


def _aplicar_formato_encabezado(
    ws: Any,
    n_cols: int,
    calc_cols: set[str] | None = None,
) -> None:
    """Aplica encabezado azul a columnas normales y verde a columnas calculadas.

    Las columnas cuyo nombre (en mayusculas) este en ``calc_cols`` reciben
    fondo verde (_CALC_HEADER_FILL) para distinguirlas visualmente de las
    columnas que vienen directamente del query SQL (fondo azul estandar).

    Args:
        ws: Hoja de trabajo de openpyxl.
        n_cols: Numero de columnas en la hoja.
        calc_cols: Conjunto de nombres de columna calculada. Si es None
            o vacio, todas las columnas reciben el encabezado azul estandar.
    """
    calc_upper: set[str] = {c.upper() for c in calc_cols} if calc_cols else set()
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        nombre = str(cell.value).upper() if cell.value else ""
        cell.font = _HEADER_FONT
        cell.fill = _CALC_HEADER_FILL if nombre in calc_upper else _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _aplicar_bordes(ws: Any, n_filas: int, n_cols: int) -> None:
    """Aplica bordes delgados a todas las celdas de datos.

    Args:
        ws: Hoja de trabajo de openpyxl.
        n_filas: Numero de filas de datos (sin encabezado).
        n_cols: Numero de columnas.
    """
    for row_idx in range(2, n_filas + 2):
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER


def _aplicar_formatos_columna(
    ws: Any,
    columnas: list[str],
    n_filas: int,
) -> None:
    """Aplica formato de moneda y fecha segun el tipo de columna.

    Args:
        ws: Hoja de trabajo de openpyxl.
        columnas: Lista de nombres de columna en orden.
        n_filas: Numero de filas de datos.
    """
    for col_idx, col_name in enumerate(columnas, start=1):
        col_upper = str(col_name).upper()
        es_moneda = (
            col_upper in COLUMNAS_MONEDA
            or any(col_upper.startswith(p) for p in _COLUMNAS_MONEDA_PREFIJOS)
        )
        if es_moneda:
            for row_idx in range(2, n_filas + 2):
                ws.cell(
                    row=row_idx, column=col_idx,
                ).number_format = "#,##0.00"
        elif col_upper in COLUMNAS_FECHA:
            for row_idx in range(2, n_filas + 2):
                ws.cell(
                    row=row_idx, column=col_idx,
                ).number_format = "DD/MM/YYYY"
        elif col_upper in COLUMNAS_PORCENTAJE:
            # Los valores de PCT_DEL_TOTAL estan en escala 0-100
            # (p.ej. 34.50 significa 34.50%). Se usa el formato con
            # simbolo literal para no que Excel no divida entre 100.
            for row_idx in range(2, n_filas + 2):
                ws.cell(
                    row=row_idx, column=col_idx,
                ).number_format = '#,##0.00"%"'



def _aplicar_bandas_alternas(
    ws: Any,
    band_data: Any,
    n_cols: int,
) -> None:
    """Aplica colores alternados por grupo de cargo/abono.

    Args:
        ws: Hoja de trabajo de openpyxl.
        band_data: Array o Series con valores 0 o 1 por fila.
        n_cols: Numero de columnas visibles.
    """
    for i, band_value in enumerate(band_data):
        row_idx = i + 2
        fill = _BAND_FILL if int(band_value) == 0 else _WHITE_FILL
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _autoajustar_ancho_columnas(ws: Any) -> None:
    """Ajusta el ancho de cada columna al contenido mas ancho.

    Recorre todas las filas de la hoja. Para celdas con formato numerico
    o de fecha estima el largo del valor formateado en lugar del valor
    crudo, evitando columnas demasiado angostas cuando el valor interno
    es un float o datetime pero se muestra como "1,234,567.89" o
    "DD/MM/YYYY".

    Ancho minimo: 10 caracteres. Ancho maximo: 60 caracteres.

    Args:
        ws: Hoja de trabajo de openpyxl.
    """
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter

        for cell in col_cells:
            if cell.value is None:
                continue

            fmt = cell.number_format or ""
            val = cell.value

            # Estimar largo segun el tipo de formato aplicado.
            if "DD/MM/YYYY" in fmt or "YYYY" in fmt:
                # Fecha formateada siempre mide 10 caracteres.
                cell_len = 10
            elif "#,##0" in fmt or "0.00" in fmt:
                # Formatear el numero como lo mostraria Excel.
                try:
                    cell_len = len(f"{float(val):,.2f}")
                except (ValueError, TypeError):
                    cell_len = len(str(val))
            elif '"%"' in fmt:
                # Porcentaje estilo 34.50%
                try:
                    cell_len = len(f"{float(val):.2f}%")
                except (ValueError, TypeError):
                    cell_len = len(str(val))
            else:
                cell_len = len(str(val))

            if cell_len > max_length:
                max_length = cell_len

        adjusted = min(max(max_length + 3, 10), 60)
        ws.column_dimensions[col_letter].width = adjusted


def _extraer_banda(df: pd.DataFrame) -> tuple[pd.DataFrame, Any]:
    """Separa la columna _BAND_GROUP del DataFrame si existe.

    Args:
        df: DataFrame que puede contener _BAND_GROUP.

    Returns:
        Tupla (DataFrame sin _BAND_GROUP, array de bandas o None).
    """
    if "_BAND_GROUP" in df.columns:
        band_data = df["_BAND_GROUP"].values.copy()
        return df.drop(columns=["_BAND_GROUP"]), band_data
    return df, None


def _escribir_hoja(
    writer: Any,
    nombre_hoja: str,
    df: pd.DataFrame,
    band_data: Any = None,
    protegida: bool = False,
    password: str = "",
    calc_cols: set[str] | None = None,
) -> None:
    """Escribe un DataFrame como hoja Excel con formato completo.

    Aplica encabezado (azul estandar o verde para columnas calculadas),
    bordes, formatos de moneda/fecha, bandas de color opcionales y
    proteccion de hoja opcional.

    Args:
        writer: ExcelWriter de pandas activo.
        nombre_hoja: Nombre de la hoja (max 31 caracteres).
        df: DataFrame a escribir sin la columna _BAND_GROUP.
        band_data: Array de bandas 0/1 para coloreo alterno. Opcional.
        protegida: Si True, aplica proteccion de solo lectura.
        password: Contrasena para la proteccion. Cadena vacia significa
            que Excel pedira dejar el campo en blanco al desproteger.
        calc_cols: Nombres de columnas calculadas que reciben encabezado
            verde. Si es None, todas reciben el encabezado azul estandar.
    """
    sheet_name = nombre_hoja[:31]
    df.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    n_filas = len(df)
    n_cols = len(df.columns)
    columnas = [str(c) for c in df.columns]

    _aplicar_formato_encabezado(ws, n_cols, calc_cols=calc_cols)
    _aplicar_bordes(ws, n_filas, n_cols)
    _aplicar_formatos_columna(ws, columnas, n_filas)

    if band_data is not None:
        _aplicar_bandas_alternas(ws, band_data, n_cols)

    _autoajustar_ancho_columnas(ws)
    ws.sheet_view.showGridLines = False

    if protegida:
        ws.protection.sheet = True
        ws.protection.password = password

    logger.info(
        "  Hoja '%s': %d filas%s",
        sheet_name,
        n_filas,
        " (protegida)" if protegida else "",
    )


def _exportar_excel(
    dataframes: dict[str, pd.DataFrame],
    nombre_base: str,
    timestamp: str,
    output_dir: Path,
    orden_hojas: list[str],
    cols_calc_por_hoja: dict[str, set[str]] | None = None,
) -> Path:
    """Exporta un conjunto de DataFrames a un solo archivo Excel.

    Itera sobre ``orden_hojas`` en orden. Las hojas vacias o ausentes
    en ``dataframes`` se omiten silenciosamente. La contrasena de cada
    hoja protegida se lee de ``SHEET_PASSWORDS`` en settings.

    Args:
        dataframes: Mapeo nombre_hoja -> DataFrame.
        nombre_base: Nombre base del archivo sin timestamp ni extension.
        timestamp: Sufijo de timestamp con formato YYYYMMDD_HHMMSS.
        output_dir: Directorio de salida; se crea si no existe.
        orden_hojas: Lista ordenada de nombres de hoja a incluir.
        cols_calc_por_hoja: Mapeo opcional nombre_hoja -> conjunto de
            nombres de columnas calculadas que reciben encabezado verde.

    Returns:
        Path al archivo .xlsx generado.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    filepath = output_dir / f"{nombre_base}_{timestamp}.xlsx"
    cols_calc_por_hoja = cols_calc_por_hoja or {}

    with pd.ExcelWriter(filepath, engine=EXCEL_ENGINE) as writer:
        for nombre_hoja in orden_hojas:
            df = dataframes.get(nombre_hoja)
            if df is None or not isinstance(df, pd.DataFrame) or df.empty:
                continue
            df, band_data = _extraer_banda(df.copy())
            protegida = nombre_hoja in PESTANAS_PROTEGIDAS
            password = SHEET_PASSWORDS.get(nombre_hoja, "")
            calc_cols = cols_calc_por_hoja.get(nombre_hoja)
            _escribir_hoja(
                writer, nombre_hoja, df, band_data, protegida, password,
                calc_cols=calc_cols,
            )

    logger.info("Excel exportado: %s", filepath)
    return filepath


# ======================================================================
# EXPORTACION — CUATRO ARCHIVOS EXCEL
# ======================================================================


def exportar_cuatro_exceles(
    cxc: dict[str, pd.DataFrame],
    auditoria: dict[str, pd.DataFrame],
    analisis: dict[str, pd.DataFrame],
    kpis: dict[str, pd.DataFrame],
    timestamp: str,
    output_dir: Path,
) -> list[Path]:
    """Exporta los resultados del pipeline a cuatro archivos Excel.

    Estructura de hojas por archivo:

    01_cxc:
        movimientos_abiertos_cxc    Cargos pendientes + abonos parciales.
        movimientos_cerrados_cxc    Cargos cobrados + todos sus abonos.
        movimientos_totales_cxc     Todas las columnas + Z-scores de
                                    IMPORTE, DELTA_RECAUDO y DELTA_MORA.
        registros_por_acreditar_cxc Anticipos sin aplicar.
        registros_cancelados        Documentos cancelados.
        registros_totales_cxc       Todos los registros sin filtros
                                    (protegida con contrasena).

    02_analisis:
        antiguedad_cartera          Distribucion por rango de dias.
        antiguedad_por_cliente_mxn  Pivot por cliente, moneda MXN.
        antiguedad_por_cliente_usd  Pivot por cliente, moneda USD.
        cartera_vencida_vs_vigente  Comparativo vencido vs vigente.
        resumen_por_vendedor        Totales y saldos por vendedor.
        resumen_por_concepto        Totales y saldos por concepto.
        resumen_ajustes             Resumen de registros por acreditar.
        resumen_cancelados          Resumen de documentos cancelados.

    03_kpis:
        kpis_resumen                DSO, CEI, Indice de Morosidad.
        kpis_concentracion          Clasificacion ABC y curva Pareto.
        kpis_limite_credito         Utilizacion del limite de credito.
        kpis_morosidad_cliente      Morosidad individual por cliente.

    Args:
        cxc: Dict con las seis hojas del reporte operativo.
        auditoria: Dict con las hojas de auditoria y calidad.
        analisis: Dict con las hojas de analisis de cartera.
        kpis: Dict con las hojas de KPIs estrategicos.
        timestamp: Sufijo de timestamp (YYYYMMDD_HHMMSS).
        output_dir: Directorio donde se guardan los archivos.

    Returns:
        Lista de Path a los cuatro archivos generados (en orden).
    """
    archivos: list[Path] = []

    logger.info("Exportando 01_cxc...")
    archivos.append(_exportar_excel(
        dataframes=cxc,
        nombre_base=EXCEL_NOMBRES["cxc"],
        timestamp=timestamp,
        output_dir=output_dir,
        orden_hojas=[
            "movimientos_abiertos_cxc",
            "movimientos_cerrados_cxc",
            "movimientos_totales_cxc",
            "registros_por_acreditar_cxc",
            "registros_cancelados_cxc",
            "registros_totales_cxc",
        ],
        cols_calc_por_hoja={
            "movimientos_totales_cxc": COLUMNAS_CALCULADAS_CXC,
        },
    ))

    logger.info("Exportando 02_analisis...")
    archivos.append(_exportar_excel(
        dataframes=analisis,
        nombre_base=EXCEL_NOMBRES["analisis"],
        timestamp=timestamp,
        output_dir=output_dir,
        orden_hojas=[
            "cartera_vencida_vs_vigente",
            "antiguedad_cartera",
            "antiguedad_por_cliente_mxn",
            "antiguedad_por_cliente_usd",
            "resumen_por_vendedor",
            "resumen_por_concepto",
            "resumen_ajustes",
            "resumen_cancelados",
        ],
    ))

    logger.info("Exportando 03_kpis...")
    archivos.append(_exportar_excel(
        dataframes=kpis,
        nombre_base=EXCEL_NOMBRES["kpis"],
        timestamp=timestamp,
        output_dir=output_dir,
        orden_hojas=[
            "kpis_resumen",
            "kpis_concentracion",
            "kpis_limite_credito",
            "kpis_morosidad_cliente",
        ],
    ))

    logger.info("Exportando 04_auditoria...")
    archivos.append(_exportar_excel(
        dataframes=auditoria,
        nombre_base=EXCEL_NOMBRES["auditoria"],
        timestamp=timestamp,
        output_dir=output_dir,
        orden_hojas=[
            "calidad_datos",
            "importes_atipicos",
            "recaudos_atipicos",
            "moras_atipicas",
            "sin_tipo_cliente",
            "sin_vendedor",
        ],
    ))

    return archivos


# ======================================================================
# PIPELINE
# ======================================================================


def run_pipeline(
    skip_audit: bool = False,
    skip_analytics: bool = False,
    skip_kpis: bool = False,
) -> int:
    """Ejecuta el pipeline completo de auditoria CxC.

    Pasos:
        1. Extraccion de datos desde Firebird.
        2. Generacion del reporte operativo y construccion del dict cxc.
        3. Auditoria y deteccion de anomalias (opcional).
        4. Analisis de cartera y antiguedad (opcional).
        5. KPIs estrategicos (opcional).
        6. Exportacion a cuatro archivos Excel.

    Args:
        skip_audit: Si True, omite el paso de auditoria.
        skip_analytics: Si True, omite el paso de analisis de cartera.
        skip_kpis: Si True, omite el paso de KPIs estrategicos.

    Returns:
        0 si el pipeline termino correctamente, 1 si hubo error fatal.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ------------------------------------------------------------------
    # 1. EXTRACCION DE DATOS
    # ------------------------------------------------------------------
    logger.info("=" * 60)
    logger.info("PASO 1: Extraccion de datos desde Firebird")
    logger.info("=" * 60)

    connector = FirebirdConnector(FIREBIRD_CONFIG)
    try:
        df = connector.execute_sql_file(SQL_FILE)
    except Exception as exc:
        logger.error("Error al extraer datos: %s", exc)
        return 1

    logger.info(
        "Datos extraidos: %d filas x %d columnas", *df.shape,
    )

    # ------------------------------------------------------------------
    # 2. REPORTE OPERATIVO + CONSTRUCCION DEL DICT CXC
    # ------------------------------------------------------------------
    logger.info("=" * 60)
    logger.info("PASO 2: Reporte operativo CxC")
    logger.info("=" * 60)

    resultado_reporte = generar_reporte_cxc(df)

    # registros_totales_cxc contiene TODOS los registros del query,
    # sin ningun filtro, con bandas de color por grupo. Es la fuente
    # de la que se derivan los dos filtros de la ultima posicion.
    registros_totales = preparar_registros_totales(df)
    registros_por_acreditar = _filtrar_por_acreditar(registros_totales)
    registros_cancelados = _filtrar_cancelados(registros_totales)

    cxc: dict[str, pd.DataFrame] = {
        "movimientos_abiertos_cxc":    resultado_reporte.get(
            "movimientos_abiertos_cxc", pd.DataFrame(),
        ),
        "movimientos_cerrados_cxc":    resultado_reporte.get(
            "movimientos_cerrados_cxc", pd.DataFrame(),
        ),
        "movimientos_totales_cxc":     resultado_reporte.get(
            "movimientos_totales_cxc", pd.DataFrame(),
        ),
        "registros_por_acreditar_cxc": registros_por_acreditar,
        "registros_cancelados_cxc":    registros_cancelados,
        "registros_totales_cxc":       registros_totales,
    }

    # ------------------------------------------------------------------
    # 3. AUDITORIA
    # ------------------------------------------------------------------
    auditoria: dict[str, pd.DataFrame] = {}

    if not skip_audit:
        logger.info("=" * 60)
        logger.info("PASO 3: Auditoria y deteccion de anomalias")
        logger.info("=" * 60)

        reporte_cxc_df = resultado_reporte.get("reporte_cxc", pd.DataFrame())
        auditor = Auditor(ANOMALIAS)
        audit_result = auditor.run_audit(df, df_reporte=reporte_cxc_df)

        logger.info("%s", "-" * 40)
        logger.info("RESUMEN DE AUDITORIA:")
        for clave, valor in audit_result.resumen.items():
            logger.info("  %-35s %s", clave, valor)
        logger.info("%s", "-" * 40)

        auditoria = {
            "calidad_datos":     audit_result.calidad_datos,
            "importes_atipicos": audit_result.importes_atipicos,
            "recaudos_atipicos": audit_result.recaudos_atipicos,
            "moras_atipicas":    audit_result.moras_atipicas,
            "sin_tipo_cliente":  audit_result.sin_tipo_cliente,
            "sin_vendedor":      audit_result.sin_vendedor,
        }

    # ------------------------------------------------------------------
    # 4. ANALISIS DE CARTERA
    # ------------------------------------------------------------------
    analisis: dict[str, pd.DataFrame] = {}

    if not skip_analytics:
        logger.info("=" * 60)
        logger.info("PASO 4: Analisis de cartera")
        logger.info("=" * 60)

        # Las vistas nombradas alimentan cada analisis con su fuente correcta.
        # movimientos_totales_cxc incluye abiertos y cerrados (sin _BAND_GROUP
        # todavia porque se paso por _agregar_zscores — lo extraemos limpio).
        vistas_analytics = {
            "movimientos_abiertos_cxc":    cxc.get(
                "movimientos_abiertos_cxc", pd.DataFrame(),
            ),
            "movimientos_totales_cxc":     cxc.get(
                "movimientos_totales_cxc", pd.DataFrame(),
            ),
            "registros_por_acreditar_cxc": cxc.get(
                "registros_por_acreditar_cxc", pd.DataFrame(),
            ),
            "registros_cancelados_cxc":    cxc.get(
                "registros_cancelados_cxc", pd.DataFrame(),
            ),
        }

        analytics_engine = Analytics(RANGOS_ANTIGUEDAD)
        analytics_result = analytics_engine.run_analytics(vistas_analytics)

        analisis = {
            "antiguedad_cartera":          analytics_result.get(
                "antiguedad_cartera", pd.DataFrame(),
            ),
            "antiguedad_por_cliente_mxn":  analytics_result.get(
                "antiguedad_por_cliente_mxn", pd.DataFrame(),
            ),
            "antiguedad_por_cliente_usd":  analytics_result.get(
                "antiguedad_por_cliente_usd", pd.DataFrame(),
            ),
            "cartera_vencida_vs_vigente":  analytics_result.get(
                "cartera_vencida_vs_vigente", pd.DataFrame(),
            ),
            "resumen_por_vendedor":        analytics_result.get(
                "resumen_por_vendedor", pd.DataFrame(),
            ),
            "resumen_por_concepto":        analytics_result.get(
                "resumen_por_concepto", pd.DataFrame(),
            ),
            "resumen_ajustes":             analytics_result.get(
                "resumen_ajustes", pd.DataFrame(),
            ),
            "resumen_cancelados":          analytics_result.get(
                "resumen_cancelados", pd.DataFrame(),
            ),
        }

    # ------------------------------------------------------------------
    # 4b. PDF DE ANALISIS
    # ------------------------------------------------------------------
    if not skip_analytics and analisis:
        logger.info("=" * 60)
        logger.info("PASO 4b: Generando PDF de analisis")
        logger.info("=" * 60)
        try:
            ts_legible = datetime.now().strftime("%Y-%m-%d %H:%M")
            pdf_path = OUTPUT_DIR / (
                f"{EXCEL_NOMBRES['pdf']}_{timestamp}.pdf"
            )
            generar_reporte_pdf(analisis, pdf_path, ts_legible)
        except Exception as exc:
            logger.warning("No se pudo generar el PDF de analisis: %s", exc)

    # ------------------------------------------------------------------
    # 5. KPIS ESTRATEGICOS
    # ------------------------------------------------------------------
    kpis: dict[str, pd.DataFrame] = {}

    if not skip_kpis:
        logger.info("=" * 60)
        logger.info("PASO 5: KPIs estrategicos")
        logger.info("=" * 60)

        kpis_result = generar_kpis(df, KPI_PERIODO_DIAS)
        kpis = {
            "kpis_resumen":           kpis_result.get(
                "kpis_resumen", pd.DataFrame(),
            ),
            "kpis_concentracion":     kpis_result.get(
                "kpis_concentracion", pd.DataFrame(),
            ),
            "kpis_limite_credito":    kpis_result.get(
                "kpis_limite_credito", pd.DataFrame(),
            ),
            "kpis_morosidad_cliente": kpis_result.get(
                "kpis_morosidad_cliente", pd.DataFrame(),
            ),
        }

    # ------------------------------------------------------------------
    # 6. EXPORTACION A CUATRO ARCHIVOS EXCEL
    # ------------------------------------------------------------------
    logger.info("=" * 60)
    logger.info("PASO 6: Exportacion a cuatro archivos Excel")
    logger.info("=" * 60)

    archivos_generados = exportar_cuatro_exceles(
        cxc=cxc,
        auditoria=auditoria,
        analisis=analisis,
        kpis=kpis,
        timestamp=timestamp,
        output_dir=OUTPUT_DIR,
    )

    logger.info("=" * 60)
    logger.info("PIPELINE COMPLETADO EXITOSAMENTE")
    logger.info(
        "%d archivos Excel generados en: %s",
        len(archivos_generados),
        OUTPUT_DIR.resolve(),
    )
    for archivo in archivos_generados:
        logger.info("  %s", archivo.name)
    logger.info("=" * 60)

    return 0


# ======================================================================
# CLI
# ======================================================================


def parse_args() -> argparse.Namespace:
    """Parsea los argumentos de linea de comandos.

    Returns:
        Namespace con flags: test_connection, skip_audit,
        skip_analytics, skip_kpis.
    """
    parser = argparse.ArgumentParser(
        description="Pipeline de auditoria CxC para Microsip",
    )
    parser.add_argument(
        "--test-connection",
        action="store_true",
        help="Solo probar conexion a Firebird.",
    )
    parser.add_argument(
        "--skip-audit",
        action="store_true",
        help="Saltar auditoria de anomalias.",
    )
    parser.add_argument(
        "--skip-analytics",
        action="store_true",
        help="Saltar analisis de cartera.",
    )
    parser.add_argument(
        "--skip-kpis",
        action="store_true",
        help="Saltar KPIs estrategicos.",
    )
    return parser.parse_args()


def main() -> int:
    """Punto de entrada principal del pipeline.

    Returns:
        0 si todo salio bien, 1 si hubo error fatal.
    """
    args = parse_args()
    logger.info("Pipeline de Auditoria CxC para Microsip")
    logger.info(
        "Fecha: %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )

    if args.test_connection:
        connector = FirebirdConnector(FIREBIRD_CONFIG)
        return 0 if connector.test_connection() else 1

    return run_pipeline(
        skip_audit=args.skip_audit,
        skip_analytics=args.skip_analytics,
        skip_kpis=args.skip_kpis,
    )


if __name__ == "__main__":
    sys.exit(main())